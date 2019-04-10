#!/usr/bin/env python
# coding=UTF-8

'''
@Author: chenxj
@Github: https://github.com/chenxj1101
@Mail: ccj799@gmail.com
@Description: 填充excel需求单
@Date: 2019-04-10 14:13:39
'''

import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import colors, Font
from prettytable import PrettyTable


control_file = sys.argv[1]
demand_path = sys.argv[2]


def get_unpass(control):
    unpass = {}
    wb = load_workbook(control)
    sheet = wb['生物信息质控报告']
    nrow = sheet.max_row
    ncol = sheet.max_column
    for i in range(4, nrow + 1):
        unpass[sheet[f'A{i}'].value] = '通过'
        for j in range(1, ncol + 1):
            if sheet.cell(row=i, column=j).fill.start_color.index == 'FFFF0000':
                unpass[sheet[f'A{i}'].value] = sheet.cell(row=i, column=j).value

    return unpass


def fill_control(file, unpass, col_samp='U', col_path='Y', col_pass='Z'):
    wb = load_workbook(file)
    sheet = wb[wb.sheetnames[0]]
    nrow = sheet.max_row
    ft = Font(color=colors.RED)
    for i in range(4, nrow + 1):
        flag = sheet[f"{col_samp}{i}"].value
        name = sheet[f'D{i}'].value
        if flag and flag in unpass.keys():
            sheet[f"{col_pass}{i}"].value = unpass[flag] if '%' not in unpass[flag] else '不通过'
            if sheet[f"{col_pass}{i}"].value == '不通过' or sheet[f"{col_pass}{i}"].value == '失控':
                sheet[f"{col_pass}{i}"].font = ft
            if sheet[f"{col_pass}{i}"].value == '不通过':
                with open(f"{demand_path}/unpass.txt", 'a', encoding='utf-8') as f:
                    print(f"{flag}\t{name}\t{unpass[flag]}", file=f)
            sheet[f"{col_path}{i}"].value = os.getcwd()
    wb.save(file)


def unpass_table(file):
    tb = PrettyTable()
    tb.field_names = ['检测编号', '姓名', '质控']
    with open(file, 'r') as f:
        for line in f:
            lines = line.strip().split('\t')
            tb.add_row(lines)
    return tb


if __name__ == '__main__':

    unpass = get_unpass(control_file)

    files = [f"{demand_path}/{f}" for f in os.listdir(demand_path) if f.endswith('xlsx')]
    for file in files:
        print(file)
        if 'BRCA' in file:
            fill_control(file, unpass, col_samp='T', col_path='X', col_pass='Y')
    else:
        fill_control(file, unpass)

    utb = unpass_table(f"{demand_path}/unpass.txt")
    print(utb)